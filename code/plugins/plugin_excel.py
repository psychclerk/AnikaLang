import os
import csv
import re

from .base_plugin import AnikaPlugin
from core.interpreter import NativeFunction
from core.errors import FMS_Error

class ExcelPlugin(AnikaPlugin):
    @staticmethod
    def _get_openpyxl():
        try:
            import openpyxl
            return openpyxl
        except ImportError:
            raise FMS_Error("Excel support requires openpyxl. Run: pip install openpyxl", error_type="Import Error")

    def register(self, env, interpreter):
        # ==========================================================================
        # EXCEL FILE MANIPULATION (Basic Read/Write)
        # ==========================================================================
        def excel_read(i, a):
            openpyxl = self._get_openpyxl()
            path = str(a[0])
            sheet_name = str(a[1]) if len(a) > 1 else None
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb[sheet_name] if sheet_name else wb.active
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) == 0:
                    wb.close()
                    return []
                headers = [str(h) if h is not None else f"col_{idx}" for idx, h in enumerate(rows[0])]
                result = []
                for row in rows[1:]:
                    row_dict = {}
                    for idx, cell in enumerate(row):
                        if idx < len(headers):
                            if cell is None: row_dict[headers[idx]] = ""
                            elif isinstance(cell, (int, float, bool)): row_dict[headers[idx]] = cell
                            else: row_dict[headers[idx]] = str(cell)
                    result.append(row_dict)
                wb.close()
                return result
            except FMS_Error: raise
            except FileNotFoundError:
                raise FMS_Error(f"Excel file not found: '{path}'", error_type="File Error")
            except Exception as e:
                raise FMS_Error(f"Failed to read Excel: {str(e)}", error_type="File Error")

        def excel_write(i, a):
            openpyxl = self._get_openpyxl()
            path = str(a[0])
            data = a[1]
            sheet_name = str(a[2]) if len(a) > 2 else "Sheet1"
            if not isinstance(data, list) or len(data) == 0:
                raise FMS_Error("EXCEL_WRITE requires a non-empty list of dictionaries", error_type="Runtime Error")
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name
                headers = list(data[0].keys())
                ws.append(headers)
                from openpyxl.styles import Font
                for cell in ws[1]: cell.font = Font(bold=True)
                for row_dict in data:
                    row_values = [row_dict.get(h, "") for h in headers]
                    ws.append(row_values)
                for col_idx, header in enumerate(headers, 1):
                    max_length = len(str(header))
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value: max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
                os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
                wb.save(path)
                wb.close()
                return "SUCCESS"
            except FMS_Error: raise
            except Exception as e:
                raise FMS_Error(f"Failed to write Excel: {str(e)}", error_type="File Error")

        def excel_append(i, a):
            openpyxl = self._get_openpyxl()
            path = str(a[0])
            data = a[1]
            sheet_name = str(a[2]) if len(a) > 2 else None
            if not isinstance(data, list) or len(data) == 0:
                raise FMS_Error("EXCEL_APPEND requires a non-empty list of dictionaries", error_type="Runtime Error")
            try:
                if os.path.exists(path): wb = openpyxl.load_workbook(path)
                else: wb = openpyxl.Workbook()
                if sheet_name:
                    if sheet_name in wb.sheetnames: ws = wb[sheet_name]
                    else:
                        ws = wb.create_sheet(sheet_name)
                        ws.append(list(data[0].keys()))
                else:
                    ws = wb.active
                for row_dict in data:
                    if ws.max_row > 0:
                        headers = [cell.value for cell in ws[1]]
                        row_values = [row_dict.get(str(h), "") for h in headers]
                    else:
                        row_values = list(row_dict.values())
                    ws.append(row_values)
                os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
                wb.save(path)
                wb.close()
                return "SUCCESS"
            except FMS_Error: raise
            except Exception as e:
                raise FMS_Error(f"Failed to append Excel: {str(e)}", error_type="File Error")

        def excel_sheets(i, a):
            openpyxl = self._get_openpyxl()
            path = str(a[0])
            try:
                wb = openpyxl.load_workbook(path, read_only=True)
                sheets = wb.sheetnames
                wb.close()
                return sheets
            except FileNotFoundError:
                raise FMS_Error(f"Excel file not found: '{path}'", error_type="File Error")
            except Exception as e:
                raise FMS_Error(f"Failed to read Excel sheets: {str(e)}", error_type="File Error")

        env.define("EXCEL_READ", NativeFunction("EXCEL_READ", -1, excel_read))
        env.define("EXCEL_WRITE", NativeFunction("EXCEL_WRITE", -1, excel_write))
        env.define("EXCEL_APPEND", NativeFunction("EXCEL_APPEND", -1, excel_append))
        env.define("EXCEL_SHEETS", NativeFunction("EXCEL_SHEETS", 1, excel_sheets))

        # ==========================================================================
        # ENHANCED XLSX MANIPULATION
        # ==========================================================================
        def xlsx_format_cell(i, a):
            openpyxl = self._get_openpyxl()
            path = str(a[0])
            sheet_name = str(a[1])
            cell_ref = str(a[2])
            bold = bool(a[3]) if len(a) > 3 else False
            font_color = str(a[4]).lstrip('#') if len(a) > 4 and a[4] else None
            bg_color = str(a[5]).lstrip('#') if len(a) > 5 and a[5] else None
            font_size = int(a[6]) if len(a) > 6 and a[6] else None
            try:
                wb = openpyxl.load_workbook(path)
                ws = wb[sheet_name] if sheet_name else wb.active
                cell = ws[cell_ref]
                from openpyxl.styles import Font, PatternFill
                if bold or font_color or font_size:
                    cell.font = Font(bold=bold, color=font_color if font_color else None, size=font_size if font_size else None)
                if bg_color:
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                wb.save(path)
                wb.close()
                return "SUCCESS"
            except Exception as e:
                raise FMS_Error(f"Failed to format cell: {str(e)}", error_type="File Error")

        def xlsx_merge_cells(i, a):
            openpyxl = self._get_openpyxl()
            path, sheet_name, cell_range = str(a[0]), str(a[1]), str(a[2])
            try:
                wb = openpyxl.load_workbook(path)
                ws = wb[sheet_name] if sheet_name else wb.active
                ws.merge_cells(cell_range)
                wb.save(path)
                wb.close()
                return "SUCCESS"
            except Exception as e:
                raise FMS_Error(f"Failed to merge cells: {str(e)}", error_type="File Error")

        def xlsx_set_column_width(i, a):
            openpyxl = self._get_openpyxl()
            path, sheet_name, col = str(a[0]), str(a[1]), str(a[2])
            width = float(a[3])
            try:
                wb = openpyxl.load_workbook(path)
                ws = wb[sheet_name] if sheet_name else wb.active
                ws.column_dimensions[col].width = width
                wb.save(path)
                wb.close()
                return "SUCCESS"
            except Exception as e:
                raise FMS_Error(f"Failed to set column width: {str(e)}", error_type="File Error")

        def xlsx_add_formula(i, a):
            openpyxl = self._get_openpyxl()
            path, sheet_name, cell_ref, formula = str(a[0]), str(a[1]), str(a[2]), str(a[3])
            try:
                wb = openpyxl.load_workbook(path)
                ws = wb[sheet_name] if sheet_name else wb.active
                ws[cell_ref] = formula
                wb.save(path)
                wb.close()
                return "SUCCESS"
            except Exception as e:
                raise FMS_Error(f"Failed to add formula: {str(e)}", error_type="File Error")

        def xlsx_add_chart(i, a):
            openpyxl = self._get_openpyxl()
            path = str(a[0])
            sheet_name = str(a[1])
            chart_type = str(a[2]).lower()
            data_range = str(a[3])
            title = str(a[4]) if len(a) > 4 else "Chart"
            anchor = str(a[5]) if len(a) > 5 else "E2"
            try:
                from openpyxl.chart import BarChart, LineChart, PieChart, Reference
                wb = openpyxl.load_workbook(path)
                ws = wb[sheet_name] if sheet_name else wb.active
                match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', data_range)
                if not match:
                    raise FMS_Error(f"Invalid data range format: '{data_range}'. Use format like 'A1:B10'", error_type="Runtime Error")
                min_col = openpyxl.utils.column_index_from_string(match.group(1))
                min_row = int(match.group(2))
                max_col = openpyxl.utils.column_index_from_string(match.group(3))
                max_row = int(match.group(4))
                data = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
                if chart_type == 'bar': chart = BarChart()
                elif chart_type == 'line': chart = LineChart()
                elif chart_type == 'pie': chart = PieChart()
                else: raise FMS_Error(f"Unknown chart type: '{chart_type}'. Use 'bar', 'line', or 'pie'", error_type="Runtime Error")
                chart.title = title
                chart.add_data(data)
                ws.add_chart(chart, anchor)
                wb.save(path)
                wb.close()
                return "SUCCESS"
            except FMS_Error: raise
            except Exception as e:
                raise FMS_Error(f"Failed to add chart: {str(e)}", error_type="File Error")

        def xlsx_get_cell(i, a):
            openpyxl = self._get_openpyxl()
            path, sheet_name, cell_ref = str(a[0]), str(a[1]), str(a[2])
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb[sheet_name] if sheet_name else wb.active
                value = ws[cell_ref].value
                wb.close()
                return value if value is not None else ""
            except Exception as e:
                raise FMS_Error(f"Failed to get cell: {str(e)}", error_type="File Error")

        def xlsx_set_cell(i, a):
            openpyxl = self._get_openpyxl()
            path, sheet_name, cell_ref, value = str(a[0]), str(a[1]), str(a[2]), a[3]
            try:
                wb = openpyxl.load_workbook(path)
                ws = wb[sheet_name] if sheet_name else wb.active
                ws[cell_ref] = value
                wb.save(path)
                wb.close()
                return "SUCCESS"
            except Exception as e:
                raise FMS_Error(f"Failed to set cell: {str(e)}", error_type="File Error")

        def xlsx_to_csv(i, a):
            openpyxl = self._get_openpyxl()
            xlsx_path, csv_path = str(a[0]), str(a[1])
            sheet_name = str(a[2]) if len(a) > 2 else None
            try:
                wb = openpyxl.load_workbook(xlsx_path, data_only=True)
                ws = wb[sheet_name] if sheet_name else wb.active
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                with open(csv_path, 'w', encoding='utf-8', newline='') as f:
                    writer = csv.writer(f)
                    for row in rows:
                        writer.writerow([c if c is not None else "" for c in row])
                return "SUCCESS"
            except Exception as e:
                raise FMS_Error(f"Failed to convert to CSV: {str(e)}", error_type="File Error")

        def csv_to_xlsx(i, a):
            openpyxl = self._get_openpyxl()
            csv_path, xlsx_path = str(a[0]), str(a[1])
            sheet_name = str(a[2]) if len(a) > 2 else "Sheet1"
            try:
                with open(csv_path, 'r', encoding='utf-8', newline='') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name
                for row in rows: ws.append(row)
                for col_idx in range(1, len(rows[0]) + 1 if rows else 1):
                    max_length = 0
                    for row in rows:
                        if col_idx - 1 < len(row):
                            max_length = max(max_length, len(str(row[col_idx - 1])))
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
                os.makedirs(os.path.dirname(xlsx_path) or '.', exist_ok=True)
                wb.save(xlsx_path)
                wb.close()
                return "SUCCESS"
            except Exception as e:
                raise FMS_Error(f"Failed to convert to XLSX: {str(e)}", error_type="File Error")

        env.define("XLSX_FORMAT_CELL", NativeFunction("XLSX_FORMAT_CELL", -1, xlsx_format_cell))
        env.define("XLSX_MERGE_CELLS", NativeFunction("XLSX_MERGE_CELLS", 3, xlsx_merge_cells))
        env.define("XLSX_SET_COLUMN_WIDTH", NativeFunction("XLSX_SET_COLUMN_WIDTH", 4, xlsx_set_column_width))
        env.define("XLSX_ADD_FORMULA", NativeFunction("XLSX_ADD_FORMULA", 4, xlsx_add_formula))
        env.define("XLSX_ADD_CHART", NativeFunction("XLSX_ADD_CHART", -1, xlsx_add_chart))
        env.define("XLSX_GET_CELL", NativeFunction("XLSX_GET_CELL", 3, xlsx_get_cell))
        env.define("XLSX_SET_CELL", NativeFunction("XLSX_SET_CELL", 4, xlsx_set_cell))
        env.define("XLSX_TO_CSV", NativeFunction("XLSX_TO_CSV", -1, xlsx_to_csv))
        env.define("CSV_TO_XLSX", NativeFunction("CSV_TO_XLSX", -1, csv_to_xlsx))